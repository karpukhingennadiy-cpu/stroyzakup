# backend/tests/test_export.py
"""P1: XLSX competitive-sheet export + PDF winner protocol + IDOR regressions."""
import io

import pytest
from django.contrib.auth import get_user_model
from django.test import override_settings
from rest_framework.test import APIClient

from apps.quotes.models import Quote, QuoteItem, RfqInvitation
from apps.requests.models import Category, Request, RequestItem, Unit
from apps.suppliers.models import Supplier

User = get_user_model()


def _client(email, password="pass"):
    User.objects.create_user(email=email, password=password, username=email)
    client = APIClient()
    r = client.post("/api/auth/login/", {"email": email, "password": password})
    client.credentials(HTTP_AUTHORIZATION=f"Bearer {r.data['access']}")
    return client


@pytest.fixture
def owner_client(db):
    return _client("exp-owner@test.com")


@pytest.fixture
def other_client(db):
    return _client("exp-other@test.com")


@pytest.fixture
def request_with_quotes(db, owner_client):
    user = User.objects.get(email="exp-owner@test.com")
    req = Request.objects.create(customer=user, code="EXP001", status="collecting_quotes",
                                 raw_text="Цемент М500 - 10 меш")
    cat, _ = Category.objects.get_or_create(slug="cement", defaults={"name": "Cement"})
    unit, _ = Unit.objects.get_or_create(code="bag", defaults={"name": "Bag", "short_name": "меш"})
    item = RequestItem.objects.create(request=req, raw_text="x", name="Цемент М500",
                                      category=cat, quantity=10, unit=unit, is_confirmed=True)
    quotes = []
    for name, price, delivery in (("BestSup", 500, 0), ("MidSup", 550, 100), ("WorstSup", 700, 50)):
        sup = Supplier.objects.create(name=name, email=f"{name.lower()}@sup.ru")
        inv = RfqInvitation.objects.create(
            request=req, supplier=sup, code=name[:8].upper(), reply_code=name.lower(),
            reply_email=f"rfq-{name.lower()}@in.example", quote_token=name.lower() * 8)
        q = Quote.objects.create(request=req, supplier=sup, invitation=inv,
                                 status="received", delivery_cost=delivery,
                                 delivery_time="3 дня", payment_terms="100% предоплата")
        QuoteItem.objects.create(quote=q, request_item=item, price=price)
        quotes.append(q)
    return req


class TestCompetitiveSheetXlsx:
    def test_xlsx_download_ok(self, owner_client, request_with_quotes):
        req = request_with_quotes
        r = owner_client.get(f"/api/quotes/competitive_sheet_xlsx/?request_id={req.id}")
        assert r.status_code == 200
        assert r["Content-Type"].startswith(
            "application/vnd.openxmlformats-officedocument.spreadsheetml")
        assert "competitive_sheet_RFQ-EXP001.xlsx" in r["Content-Disposition"]

    def test_xlsx_content_and_best_highlight(self, owner_client, request_with_quotes):
        from openpyxl import load_workbook

        req = request_with_quotes
        r = owner_client.get(f"/api/quotes/competitive_sheet_xlsx/?request_id={req.id}")
        wb = load_workbook(io.BytesIO(b"".join(r.streaming_content)
                                      if hasattr(r, "streaming_content") else r.content))
        ws = wb.active
        # Header row
        assert ws["B6"].value == "Поставщик"
        # Rows sorted by grand total -> BestSup first, marked with ★
        assert "BestSup" in ws["B7"].value
        assert ws["B7"].value.startswith("★")
        # Total is a formula (materials + delivery), not a hardcoded value
        assert str(ws["G7"].value).startswith("=")
        # Best row is highlighted
        assert ws["B7"].fill.start_color.rgb == "00E6F0FA"
        # 3 supplier rows present
        assert "MidSup" in ws["B8"].value
        assert "WorstSup" in ws["B9"].value

    def test_xlsx_other_users_request_404(self, other_client, request_with_quotes):
        req = request_with_quotes
        r = other_client.get(f"/api/quotes/competitive_sheet_xlsx/?request_id={req.id}")
        assert r.status_code == 404

    def test_xlsx_no_quotes_still_works(self, owner_client, db):
        user = User.objects.get(email="exp-owner@test.com")
        req = Request.objects.create(customer=user, code="EXP002", status="matched")
        r = owner_client.get(f"/api/quotes/competitive_sheet_xlsx/?request_id={req.id}")
        assert r.status_code == 200
        assert len(r.content) > 1000  # a real workbook, not an error page


class TestWinnerProtocolPdf:
    def test_pdf_download_ok(self, owner_client, request_with_quotes):
        req = request_with_quotes
        r = owner_client.get(f"/api/quotes/winner_protocol_pdf/?request_id={req.id}")
        assert r.status_code == 200
        assert r["Content-Type"] == "application/pdf"
        assert b"%PDF" in r.content[:8]
        assert len(r.content) > 2000

    def test_pdf_other_users_request_404(self, other_client, request_with_quotes):
        req = request_with_quotes
        r = other_client.get(f"/api/quotes/winner_protocol_pdf/?request_id={req.id}")
        assert r.status_code == 404

    def test_pdf_no_quotes_still_works(self, owner_client, db):
        user = User.objects.get(email="exp-owner@test.com")
        req = Request.objects.create(customer=user, code="EXP003", status="matched")
        r = owner_client.get(f"/api/quotes/winner_protocol_pdf/?request_id={req.id}")
        assert r.status_code == 200
        assert b"%PDF" in r.content[:8]

    def test_pdf_requires_request_id(self, owner_client):
        r = owner_client.get("/api/quotes/winner_protocol_pdf/")
        assert r.status_code == 400

    def test_pdf_delivery_term_has_unit(self, owner_client, request_with_quotes):
        """Regression (protocol_ZUABCR): a bare numeric delivery_time must be
        rendered with its unit, declined ('5' -> '5 дней')."""
        from pypdf import PdfReader

        req = request_with_quotes
        best_quote = Quote.objects.get(supplier__name="BestSup")
        best_quote.delivery_time = "5"
        best_quote.save(update_fields=["delivery_time"])

        r = owner_client.get(f"/api/quotes/winner_protocol_pdf/?request_id={req.id}")
        assert r.status_code == 200
        text = "".join(
            page.extract_text() for page in PdfReader(io.BytesIO(r.content)).pages
        )
        assert "5 дней" in text

    def test_format_delivery_term_plural_rules(self, db):
        from apps.quotes.exporters import format_delivery_term

        assert format_delivery_term("1") == "1 день"
        assert format_delivery_term("2") == "2 дня"
        assert format_delivery_term("5") == "5 дней"
        assert format_delivery_term("11") == "11 дней"
        assert format_delivery_term("21") == "21 день"
        assert format_delivery_term(3) == "3 дня"
        assert format_delivery_term("3 дня") == "3 дня"  # already has a unit
        assert format_delivery_term("") == "—"
        assert format_delivery_term(None) == "—"


class TestUpdateItemIdorRegression:
    """update_item used to accept any request id — a user could edit items
    of another customer's request."""

    def test_update_item_other_users_request_404(self, other_client, request_with_quotes):
        req = request_with_quotes
        item = req.items.first()
        r = other_client.post(
            f"/api/requests/{req.id}/update_item/",
            {"item_id": item.id, "is_confirmed": True}, format="json")
        assert r.status_code == 404

    def test_update_item_owner_ok(self, owner_client, request_with_quotes):
        req = request_with_quotes
        item = req.items.first()
        r = owner_client.post(
            f"/api/requests/{req.id}/update_item/",
            {"item_id": item.id, "is_confirmed": True, "brand": "Holcim"}, format="json")
        assert r.status_code == 200
        item.refresh_from_db()
        assert item.brand == "Holcim"


class TestGenericWebhookSecret:
    """generic_inbound_webhook must reject requests without the shared secret
    when INBOUND_GENERIC_WEBHOOK_SECRET is configured."""

    @override_settings(INBOUND_GENERIC_WEBHOOK_SECRET="s3cret")
    def test_webhook_without_secret_403(self, db):
        client = APIClient()
        r = client.post("/api/emails/webhook/inbound/",
                        {"subject": "x"}, format="json")
        assert r.status_code == 403

    @override_settings(INBOUND_GENERIC_WEBHOOK_SECRET="s3cret")
    def test_webhook_with_secret_200(self, db):
        client = APIClient()
        r = client.post("/api/emails/webhook/inbound/",
                        {"subject": "x"}, format="json",
                        HTTP_X_WEBHOOK_SECRET="s3cret")
        assert r.status_code == 200
