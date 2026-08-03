"""P1: competitive-sheet export to XLSX and winner-selection protocol to PDF.

Shared data builder (get_competitive_rows) is used by the API view and both
exporters so the numbers always match the JSON competitive sheet.
"""
import io
import logging
from decimal import Decimal

from django.conf import settings
from django.utils import timezone

logger = logging.getLogger(__name__)

MONEY_FMT = '#,##0.00 "₽"'


def _plural_days(n: int) -> str:
    """день/дня/дней by Russian plural rules."""
    if 11 <= n % 100 <= 14:
        return "дней"
    return {1: "день", 2: "дня", 3: "дня", 4: "дня"}.get(n % 10, "дней")


def format_delivery_term(value) -> str:
    """Human-readable delivery term: a bare number gets its unit declined
    ('3' -> '3 дня'); values that already contain text pass through."""
    if value is None:
        return "—"
    text = str(value).strip()
    if not text:
        return "—"
    try:
        num = float(text.replace(",", "."))
    except ValueError:
        return text  # already descriptive, e.g. '3 дня', 'на следующей неделе'
    if num.is_integer():
        return f"{int(num)} {_plural_days(int(num))}"
    return f"{num:g} дня"


def get_competitive_rows(request_obj) -> list[dict]:
    """Quotes of a request ranked by grand total (materials + delivery)."""
    from .models import Quote

    quotes = (
        Quote.objects.filter(request=request_obj, status__in=["received", "valid"])
        .select_related("supplier")
        .prefetch_related("items__request_item")
    )
    rows = []
    for quote in quotes:
        total = sum(
            (qi.price * qi.request_item.quantity for qi in quote.items.all()),
            Decimal("0"),
        )
        delivery = quote.delivery_cost or Decimal("0")
        rows.append({
            "supplier_id": quote.supplier_id,
            "supplier_name": quote.supplier.name,
            "materials_total": float(total),
            "delivery": float(delivery),
            "grand_total": float(total + delivery),
            "payment_terms": quote.payment_terms,
            "delivery_time": quote.delivery_time,
            "valid_until": quote.valid_until,
        })
    rows.sort(key=lambda r: r["grand_total"])
    return rows


# === XLSX ===

def build_competitive_sheet_xlsx(request_obj) -> bytes:
    """Competitive sheet as an .xlsx workbook; the best offer is highlighted."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    rows = get_competitive_rows(request_obj)

    wb = Workbook()
    ws = wb.active
    ws.title = "Конкурентный лист"
    ws.sheet_view.showGridLines = False

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
    best_fill = PatternFill(start_color="E6F0FA", end_color="E6F0FA", fill_type="solid")
    best_font = Font(bold=True, color="0066CC")
    title_font = Font(size=14, bold=True)

    ws.row_dimensions[2].height = 24
    ws.merge_cells("B2:G2")
    ws["B2"] = f"Конкурентный лист — заявка RFQ-{request_obj.code}"
    ws["B2"].font = title_font
    ws["B2"].alignment = Alignment(horizontal="left", vertical="center")

    address = request_obj.address.address if request_obj.address else "не указан"
    ws["B3"] = f"Адрес доставки: {address}"
    ws["B3"].font = Font(color="666666")
    ws["B4"] = f"Сформировано: {timezone.now().strftime('%d.%m.%Y %H:%M')}"
    ws["B4"].font = Font(color="666666")

    headers = ["Поставщик", "Цена материалов, ₽", "Доставка, ₽",
               "Сроки поставки", "Условия оплаты", "Итого, ₽"]
    header_row = 6
    for col, title in enumerate(headers, start=2):  # B..G
        cell = ws.cell(row=header_row, column=col, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    best_total = rows[0]["grand_total"] if rows else None
    first_data_row = header_row + 1
    for idx, row in enumerate(rows):
        r = first_data_row + idx
        is_best = row["grand_total"] == best_total
        ws.cell(row=r, column=2, value=row["supplier_name"])
        ws.cell(row=r, column=3, value=row["materials_total"]).number_format = MONEY_FMT
        ws.cell(row=r, column=4, value=row["delivery"]).number_format = MONEY_FMT
        ws.cell(row=r, column=5, value=row["delivery_time"] or "—")
        ws.cell(row=r, column=6, value=row["payment_terms"] or "—")
        # Keep the total formula-driven (materials + delivery), not a static value
        total_cell = ws.cell(row=r, column=7, value=f"=C{r}+D{r}")
        total_cell.number_format = MONEY_FMT
        if is_best:
            for col in range(2, 8):
                ws.cell(row=r, column=col).fill = best_fill
                ws.cell(row=r, column=col).font = best_font
            ws.cell(row=r, column=2, value=f"★ {row['supplier_name']}")

    if not rows:
        ws.cell(row=first_data_row, column=2,
                value="По заявке пока не получено ни одного КП.")

    widths = [34, 18, 14, 18, 22, 16]
    for i, width in enumerate(widths, start=2):
        ws.column_dimensions[get_column_letter(i)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# === PDF ===

_FONT_REGISTERED: tuple[str, str] | None = None


def _register_fonts() -> tuple[str, str]:
    """Register a Cyrillic-capable TTF. Bundled DejaVu is preferred so the
    protocol renders identically in dev (Windows) and prod (Linux/Docker)."""
    global _FONT_REGISTERED
    if _FONT_REGISTERED:
        return _FONT_REGISTERED
    from pathlib import Path
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    fonts_dir = Path(settings.BASE_DIR) / "assets" / "fonts"
    candidates = [
        (fonts_dir / "DejaVuSans.ttf", fonts_dir / "DejaVuSans-Bold.ttf"),
        (Path("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"),
         Path("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf")),
        (Path("C:/Windows/Fonts/arial.ttf"), Path("C:/Windows/Fonts/arialbd.ttf")),
    ]
    for regular, bold in candidates:
        if regular.exists():
            pdfmetrics.registerFont(TTFont("ExportFont", str(regular)))
            if bold.exists():
                pdfmetrics.registerFont(TTFont("ExportFont-Bold", str(bold)))
            else:
                pdfmetrics.registerFont(TTFont("ExportFont-Bold", str(regular)))
            _FONT_REGISTERED = ("ExportFont", "ExportFont-Bold")
            return _FONT_REGISTERED
    raise RuntimeError("No Cyrillic-capable TTF font found for PDF export")


def build_winner_protocol_pdf(request_obj) -> bytes:
    """Winner-selection protocol (протокол выбора победителя) as PDF bytes."""
    from django.utils import timezone
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
    )

    font, font_bold = _register_fonts()
    rows = get_competitive_rows(request_obj)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=20 * mm, rightMargin=20 * mm,
        topMargin=20 * mm, bottomMargin=20 * mm,
        title=f"Протокол выбора победителя RFQ-{request_obj.code}",
    )

    title_style = ParagraphStyle("title", fontName=font_bold, fontSize=16, spaceAfter=6)
    h_style = ParagraphStyle("h", fontName=font_bold, fontSize=12, spaceBefore=12, spaceAfter=4)
    body_style = ParagraphStyle("body", fontName=font, fontSize=10, leading=14)
    small_style = ParagraphStyle("small", fontName=font, fontSize=9, textColor=colors.HexColor("#666666"))

    def fmt_money(value: float) -> str:
        return f"{value:,.2f} ₽".replace(",", " ")

    story = [
        Paragraph("ПРОТОКОЛ ВЫБОРА ПОБЕДИТЕЛЯ", title_style),
        Paragraph(f"Заявка № RFQ-{request_obj.code}", h_style),
        Paragraph(
            f"Дата составления: {timezone.now().strftime('%d.%m.%Y')}<br/>"
            f"Заказчик: {request_obj.customer.email}<br/>"
            f"Адрес доставки: {request_obj.address.address if request_obj.address else 'не указан'}<br/>"
            f"Получено коммерческих предложений: {len(rows)}",
            body_style,
        ),
        Spacer(1, 6 * mm),
    ]

    if rows:
        # Wrap text cells in Paragraphs so long values wrap instead of
        # overflowing the narrow columns
        cell_style = ParagraphStyle("cell", fontName=font, fontSize=8.5, leading=11)
        cell_style_b = ParagraphStyle("cellb", fontName=font_bold, fontSize=8.5, leading=11)
        head_style = ParagraphStyle("head", fontName=font_bold, fontSize=8.5,
                                    leading=10, textColor=colors.white)
        headers = ["№", "Поставщик", "Материалы, ₽", "Доставка, ₽",
                   "Срок поставки", "Оплата", "Итого, ₽"]
        table_data = [[Paragraph(h, head_style) for h in headers]]
        for i, row in enumerate(rows, 1):
            cstyle = cell_style_b if i == 1 else cell_style
            table_data.append([
                str(i), Paragraph(row["supplier_name"], cstyle),
                fmt_money(row["materials_total"]), fmt_money(row["delivery"]),
                Paragraph(format_delivery_term(row["delivery_time"]), cstyle),
                Paragraph(row["payment_terms"] or "—", cstyle),
                fmt_money(row["grand_total"]),
            ])
        col_widths = [8 * mm, 46 * mm, 26 * mm, 22 * mm, 22 * mm, 22 * mm, 24 * mm]
        table = Table(table_data, colWidths=col_widths, repeatRows=1)
        style = [
            ("FONTNAME", (0, 0), (-1, 0), font_bold),
            ("FONTNAME", (0, 1), (-1, -1), font),
            ("FONTSIZE", (0, 0), (-1, -1), 8.5),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#333333")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D0D0D0")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F5F5")]),
            # Highlight the winning row
            ("BACKGROUND", (0, 1), (-1, 1), colors.HexColor("#E6F0FA")),
            ("FONTNAME", (0, 1), (-1, 1), font_bold),
        ]
        table.setStyle(TableStyle(style))
        story += [Paragraph("Поступившие предложения", h_style), table, Spacer(1, 6 * mm)]

        best = rows[0]
        story += [
            Paragraph("Решение", h_style),
            Paragraph(
                f"Победителем признан поставщик <b>{best['supplier_name']}</b> "
                f"с наименьшей итоговой стоимостью <b>{fmt_money(best['grand_total'])}</b> "
                f"(материалы — {fmt_money(best['materials_total'])}, "
                f"доставка — {fmt_money(best['delivery'])}).<br/>"
                f"Срок поставки: {format_delivery_term(best['delivery_time'])}.<br/>"
                f"Условия оплаты: {best['payment_terms'] or '—'}.",
                body_style,
            ),
            Paragraph(
                "Основание: наименьшая итоговая стоимость среди полученных "
                "коммерческих предложений, соответствующих спецификации заявки.",
                small_style,
            ),
        ]
    else:
        story.append(Paragraph(
            "На момент составления протокола коммерческие предложения не получены. "
            "Выбор победителя невозможен.", body_style))

    story += [
        Spacer(1, 14 * mm),
        Paragraph("Заказчик: ____________________ / ____________________ /", body_style),
        Spacer(1, 3 * mm),
        Paragraph("Дата: «____» __________ 20___ г.", body_style),
    ]
    doc.build(story)
    return buf.getvalue()
